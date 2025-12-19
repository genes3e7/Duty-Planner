"""
data_manager.py

Handles all File Input/Output operations.
Implements defensive programming strategies to handle missing files,
corrupt JSON, and locked Excel files gracefully.
"""

import json
import os
import re
import logging
from typing import Dict, List, Tuple, Any, Union, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore

import constants as C


class DataManager:
    """Static utility class for file handling operations."""

    @staticmethod
    def load_config(filepath: str = C.CONFIG_FILE) -> Dict[str, Any]:
        """Loads configuration or returns defaults."""
        if not os.path.exists(filepath):
            logging.warning(f"Config file '{filepath}' not found. Loading defaults.")
            return C.DEFAULT_CONFIG_TEMPLATE.copy()

        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data: Dict[str, Any] = json.load(f)
                merged = C.DEFAULT_CONFIG_TEMPLATE.copy()
                merged.update(data)
                return merged
        except Exception as e:
            logging.error(f"Config load error: {e}. Loading defaults.")
            return C.DEFAULT_CONFIG_TEMPLATE.copy()

    @staticmethod
    def save_config(data: Dict[str, Any], filepath: str = C.CONFIG_FILE) -> None:
        """Saves configuration to JSON."""
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4)
            logging.info(f"Configuration saved to {filepath}.")
        except IOError as e:
            logging.error(f"Failed to save config: {e}")
            raise IOError(f"Could not save settings: {str(e)}")

    @staticmethod
    def sanitize_filename(filename: str) -> str:
        """Removes illegal characters from filenames."""
        return re.sub(r'[\\/*?:"<>|]', "", filename)

    @staticmethod
    def load_previous_balance(filepath: str) -> Dict[str, float]:
        """Reads previous month's balance from JSON or Excel."""
        if not filepath or not os.path.exists(filepath):
            raise FileNotFoundError(f"File not found: {filepath}")

        try:
            # JSON Handling
            if filepath.lower().endswith('.json'):
                with open(filepath, 'r', encoding='utf-8') as f:
                    data: Dict[str, float] = json.load(f)
                    return data

            # Excel Handling
            elif filepath.lower().endswith(('.xlsx', '.xls')):
                try:
                    df = pd.read_excel(filepath)
                except Exception as e:
                    raise ValueError(f"Invalid Excel file: {e}")

                cols: List[str] = [str(c).lower() for c in df.columns]
                
                name_idx = next((i for i, c in enumerate(cols) if 'name' in c), None)
                bal_idx = next((i for i, c in enumerate(cols) 
                                if any(x in c for x in ['carry', 'bal', 'roll'])), None)

                if name_idx is None or bal_idx is None:
                    raise ValueError("Excel must have 'Name' and 'Carry Over' columns.")

                name_col = df.columns[name_idx]
                bal_col = df.columns[bal_idx]

                df_clean = df[[name_col, bal_col]].dropna()
                result: Dict[str, float] = {}
                
                for _, row in df_clean.iterrows():
                    try:
                        name_key = str(row[name_col]).strip()
                        val = float(row[bal_col])
                        result[name_key] = val
                    except (ValueError, TypeError):
                        continue 

                if not result:
                    raise ValueError("File parsed, but no valid data rows found.")
                
                return result
            else:
                raise ValueError("Unsupported format. Use .xlsx or .json")

        except Exception as e:
            logging.error(f"Error loading balance file: {e}")
            raise e

    @staticmethod
    def export_schedule(
        schedule_data: Dict[Tuple[str, int], str],
        point_summary: List[Dict[str, Union[str, float]]],
        config: Dict[str, Any],
        leaves: List[Tuple[str, int]],
        save_path: str
    ) -> str:
        """Exports the generated schedule to Excel with formatting."""
        try:
            days_in_month = max([k[1] for k in schedule_data.keys()]) if schedule_data else 30
            names: List[str] = sorted(config.get('personnel', []))

            wb = Workbook()
            ws = wb.active
            ws.title = C.EXCEL_SHEET_TITLE

            # Build Headers
            headers: List[Any] = C.EXCEL_HEADERS_STATIC + \
                                 [d for d in range(1, days_in_month + 1)] + \
                                 C.EXCEL_HEADERS_SUFFIX
            ws.append(headers)

            pt_map = {str(item['Name']): item for item in point_summary}

            for name in names:
                row: List[Any] = [name]
                for day in range(1, days_in_month + 1):
                    if (name, day) in leaves:
                        row.append("X")
                    else:
                        row.append(schedule_data.get((name, day), ""))

                p_data = pt_map.get(name, {'Brought Fwd': 0.0, 'Month Pts': 0.0, 'Carry Over': 0.0})
                row.append(float(p_data['Brought Fwd'])) # type: ignore
                row.append(float(p_data['Month Pts']))   # type: ignore
                row.append(float(p_data['Carry Over']))  # type: ignore
                ws.append(row)

            # Styling
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            fill_header = PatternFill("solid", fgColor=C.COLOR_HEADER_BG)
            fill_x = PatternFill("solid", fgColor=C.COLOR_CONSTRAINT_BG)

            for excel_row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                          min_col=1, max_col=ws.max_column):
                for cell in excel_row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    if cell.row == 1:
                        cell.font = Font(bold=True)
                        cell.fill = fill_header
                    if cell.value == "X":
                        cell.fill = fill_x

            wb.save(save_path)
            logging.info(f"Export successful: {save_path}")
            return f"Successfully saved to:\n{save_path}"

        except PermissionError:
            error_msg = "File is open in Excel. Close it and try again."
            logging.error(error_msg)
            raise PermissionError(error_msg)
        except Exception as e:
            logging.error(f"Export failed: {e}")
            raise Exception(f"Failed to save Excel: {str(e)}")
