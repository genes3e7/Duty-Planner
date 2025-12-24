"""
data_manager.py

Handles File I/O operations (JSON Config, Excel Import/Export).
"""

import json
import logging
import os
from typing import Any, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app import constants as C
from app.models.config import AppConfig


class DataManager:
    """Static Utility Class for Data persistence."""

    @staticmethod
    def load_config(filepath: str = C.CONFIG_FILE) -> AppConfig:
        """Loads JSON config into AppConfig object."""
        if not os.path.exists(filepath):
            return AppConfig.default()
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
                return AppConfig.from_dict(data)
        except Exception as e:
            logging.error(f"Config load error: {e}")
            return AppConfig.default()

    @staticmethod
    def save_config(config: AppConfig, filepath: str = C.CONFIG_FILE) -> None:
        """Saves AppConfig object to JSON."""
        try:
            data = config.to_dict()
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=4)
        except IOError as e:
            logging.error(f"Config save failed: {e}")
            raise e

    @staticmethod
    def load_previous_balance(filepath: str) -> Dict[str, float]:
        """
        Imports balances from Excel.
        Uses fuzzy matching for column headers.
        """
        if not filepath or not os.path.exists(filepath):
            return {}
        try:
            df = pd.read_excel(filepath)
            cols = [str(c).lower().strip() for c in df.columns]

            name_idx = next((i for i, c in enumerate(cols) if "name" in c), None)
            # Fix: Broken down to satisfy line length limits
            bal_idx = next(
                (
                    i
                    for i, c in enumerate(cols)
                    if any(x in c for x in ["carry", "bal", "roll"])
                ),
                None,
            )

            if name_idx is None or bal_idx is None:
                return {}

            result = {}
            for _, row in df.iterrows():
                try:
                    name = str(row[df.columns[name_idx]]).strip()
                    val = float(row[df.columns[bal_idx]])
                    if name:
                        result[name] = val
                except Exception:
                    continue
            return result
        except Exception:
            return {}

    @staticmethod
    def export_schedule(
        schedule_data: Dict[Any, str],
        point_summary: List[Dict],
        config: AppConfig,
        save_path: str,
    ) -> None:
        """Exports grid and summary to formatted Excel."""
        days = max([k[1] for k in schedule_data.keys()]) if schedule_data else 30

        wb = Workbook()
        ws = wb.active
        ws.title = C.EXCEL_SHEET_TITLE

        ws.append(
            C.EXCEL_HEADERS_STATIC
            + [d for d in range(1, days + 1)]
            + C.EXCEL_HEADERS_SUFFIX
        )

        pt_map = {str(i["Name"]): i for i in point_summary}

        for name in sorted(config.personnel):
            row = [name]
            for d in range(1, days + 1):
                row.append(schedule_data.get((name, d), ""))

            p = pt_map.get(
                name, {"Brought Fwd": 0.0, "Month Pts": 0.0, "Carry Over": 0.0}
            )
            row.extend([p["Brought Fwd"], p["Month Pts"], p["Carry Over"]])
            ws.append(row)

        # Styles
        fill_header = PatternFill("solid", fgColor=C.COLOR_HEADER_BG.replace("#", ""))
        fill_x = PatternFill("solid", fgColor=C.COLOR_CONSTRAINT_BG.replace("#", ""))
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
                if cell.row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = fill_header
                if cell.value == C.ShiftType.LEAVE.value:
                    cell.fill = fill_x

        wb.save(save_path)
