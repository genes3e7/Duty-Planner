"""
app/logic.py

This module serves as the 'Controller' in the MVC pattern.
It bridges the gap between the Streamlit UI (View) and the Data/Scheduler (Model).
It handles data transformation, safe parsing, and orchestrating the solving process.
"""

import io
import logging
import re
from typing import Any, Dict, List, Optional, Tuple

import holidays
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app import constants as C
from app.core.scheduler import DutySchedulerEngine, SolverRequest
from app.models.config import AppConfig
from app.utils.helpers import get_base_shift_type, get_shift_name

logger = logging.getLogger(__name__)


def get_day_num(col_name: str) -> int:
    """
    Safely extracts the day integer from a column string (e.g., 'D1' -> 1).
    Returns 0 if the format does not match.
    """
    match = re.match(r"^D(\d+)$", str(col_name))
    if match:
        return int(match.group(1))
    return 0


def get_holidays(year: int, country_code: str = "SG") -> holidays.HolidayBase:
    """
    Returns the holiday object for the given country and year.
    Defaults to Singapore (SG) if the provided code is invalid.
    """
    try:
        return holidays.country_holidays(country_code, years=year)
    except Exception as e:
        logger.warning(f"Could not load holidays for '{country_code}': {e}. Fallback to SG.")
        return holidays.SG(years=year)


def generate_empty_schedule(
    year: int, month: int, personnel: List[str], country_code: str = "SG"
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Creates the initial empty DataFrames for the Roster and Day Configuration.

    Args:
        year (int): The selected year.
        month (int): The selected month.
        personnel (List[str]): List of staff names.
        country_code (str): Country code for holiday generation.

    Returns:
        Tuple[pd.DataFrame, pd.DataFrame]: (Roster DataFrame, Day Config DataFrame).
    """
    try:
        period = pd.Period(f"{year}-{month}")
        num_days = period.days_in_month
    except ValueError:
        logger.warning(f"Invalid year/month ({year}/{month}), defaulting to 30 days")
        num_days = 30

    country_holidays = get_holidays(year, country_code)
    day_data = []
    day_columns = [f"D{d}" for d in range(1, num_days + 1)]

    for d in range(1, num_days + 1):
        try:
            dt = pd.Timestamp(year=year, month=month, day=d)
        except ValueError as e:
            raise ValueError(f"Invalid date generated: {year}-{month}-{d}") from e

        is_ph = dt in country_holidays
        mode = C.ScheduleMode.FULL_24H.value if is_ph else C.ScheduleMode.SHIFT.value

        day_data.append(
            {
                "Day": d,
                "Date": dt.strftime("%a %d"),
                "Active": True,
                "Mode": mode,
                "Is_PH": is_ph,
                "Is_Weekend": dt.dayofweek >= 5,
            }
        )

    df_days = pd.DataFrame(day_data)
    df_days.set_index("Day", inplace=True)

    df_roster = pd.DataFrame(index=personnel, columns=day_columns)
    df_roster[:] = ""

    return df_roster, df_days


def clear_schedule(df_roster: Optional[pd.DataFrame], clear_constraints: bool = False) -> Optional[pd.DataFrame]:
    """
    Clears data from the roster grid.

    Args:
        df_roster (pd.DataFrame): The current roster.
        clear_constraints (bool): If True, clears everything. If False, keeps 'X' (unavailable).

    Returns:
        Optional[pd.DataFrame]: The cleared dataframe.
    """
    if df_roster is None:
        return None

    df = df_roster.copy(deep=True)

    if clear_constraints:
        df[:] = ""
    else:

        def keep_x(val: Any) -> str:
            """Preserves 'X' markers in the roster while clearing other values.

            Args:
                val: The cell value to process.

            Returns:
                "X" if the input is an "X" marker, otherwise an empty string.
            """
            if isinstance(val, str) and val.strip().upper() == "X":
                return "X"
            return ""

        for col in df.columns:
            df[col] = df[col].apply(keep_x)

    return df


def apply_imported_constraints(
    df_roster: pd.DataFrame, imported_data: Dict[str, Dict[int, str]]
) -> Optional[pd.DataFrame]:
    """
    Updates the roster dataframe with imported values.

    Args:
        df_roster: The existing pandas DataFrame for the roster.
        imported_data: Dictionary {Name: {DayInt: Value}} from DataManager.load_constraints.

    Returns:
        Optional[pd.DataFrame]: The updated DataFrame, or None if input roster is None.
    """
    if df_roster is None or not imported_data:
        return df_roster

    df = df_roster.copy()

    for name, day_map in imported_data.items():
        if name in df.index:
            for day_num, val in day_map.items():
                col_name = f"D{day_num}"
                if col_name in df.columns:
                    df.at[name, col_name] = val

    return df


def prepare_solver_request(
    year: int, month: int, df_roster: pd.DataFrame, df_days: pd.DataFrame, config: AppConfig
) -> SolverRequest:
    """
    Transforms the UI DataFrames into a clean SolverRequest object.
    Includes calculation of specific point weights for every day/shift.
    """
    fixed_assignments = {}
    day_modes = {}
    inactive_days = []
    shift_weights = {}

    valid_staff = set(config.personnel)
    country_holidays = get_holidays(year, config.country_code)

    # 1. Parse Day Configuration & Calculate Weights
    for day_num, row in df_days.iterrows():
        if not row["Active"]:
            inactive_days.append(day_num)
        day_modes[day_num] = row["Mode"]

        try:
            current_date = pd.Timestamp(year=year, month=month, day=day_num)

            # --- Calculate weights for ALL possible teams ---
            # 1. Active Teams (AM, PM, 24H)
            for t in range(1, config.constraints.num_active_teams + 1):
                for base_shift in ["AM", "PM", "24H"]:
                    shift_name = get_shift_name(base_shift, t)
                    # Use base_shift to look up point value (AM_2 gets AM points)
                    w = config.points.calculate_score(
                        current_date, shift_name, scale=C.SCORE_SCALE_FACTOR, holidays_obj=country_holidays
                    )
                    shift_weights[(day_num, shift_name)] = w

            # 2. Standby Teams (S/B)
            for t in range(1, config.constraints.num_standby_teams + 1):
                shift_name = get_shift_name("S/B", t)
                w = config.points.calculate_score(
                    current_date, shift_name, scale=C.SCORE_SCALE_FACTOR, holidays_obj=country_holidays
                )
                shift_weights[(day_num, shift_name)] = w
            # ------------------------------------------------

        except ValueError as e:
            logger.error(f"Invalid date configuration for Year={year}, Month={month}, Day={day_num}: {e}")
            raise ValueError(f"Invalid date encountered: {year}-{month}-{day_num}") from e

    # 2. Parse Roster Grid (Fixed Constraints)
    for person in df_roster.index:
        if person not in valid_staff:
            continue

        for day_col in df_roster.columns:
            val = df_roster.at[person, day_col]
            if val:
                day_idx = get_day_num(day_col)
                if day_idx > 0:
                    fixed_assignments[(person, day_idx)] = val

    return SolverRequest(
        staff_ids=config.personnel,
        year=year,
        month=month,
        fixed_assignments=fixed_assignments,
        day_modes=day_modes,
        inactive_days=inactive_days,
        shift_weights=shift_weights,
    )


def run_solver(
    year: int,
    month: int,
    df_roster: pd.DataFrame,
    df_days: pd.DataFrame,
    config: AppConfig,
    prev_balance: Dict[str, float],
) -> Optional[Tuple[Dict[Tuple[str, int], str], int]]:
    """
    Orchestrates the solving process.

    Returns:
        Optional[Tuple[Dict, int]]: (Schedule Dictionary, Solver Status Code) or None on failure.
    """
    try:
        req = prepare_solver_request(year, month, df_roster, df_days, config)
        engine = DutySchedulerEngine(config, prev_balance, req)
        engine.build_model()
        return engine.solve()
    except (ValueError, RuntimeError) as e:
        logger.error(f"Solver execution failed: {e}")
        return None
    except Exception as e:
        logger.exception(f"Unexpected error in solver: {e}")
        return None


def calculate_stats(
    df_roster: pd.DataFrame,
    df_days: pd.DataFrame,
    config: AppConfig,
    prev_balance: Dict[str, float],
    manual_adjustments: Optional[Dict[str, float]] = None,
) -> pd.DataFrame:
    """
    Calculates point statistics for the current roster state.
    Computes 'Month Pts' based on assignments and 'Carry Over' based on previous balance.

    Args:
        manual_adjustments: Optional dictionary of {Name: AdjustmentValue}
                            to add/subtract from monthly total.
    """
    summary = []
    raw_carry_overs = []
    country_holidays = get_holidays(config.year, config.country_code)

    if manual_adjustments is None:
        manual_adjustments = {}

    for person in config.personnel:
        bf = prev_balance.get(person, 0.0)
        adj = manual_adjustments.get(person, 0.0)

        roster_pts = 0.0

        if df_roster is not None and person in df_roster.index:
            for day_col in df_roster.columns:
                day_idx = get_day_num(day_col)
                if day_idx <= 0:
                    continue

                if day_idx not in df_days.index or not df_days.loc[day_idx, "Active"]:
                    continue

                val = df_roster.at[person, day_col]
                if val:
                    # Logic updated to handle suffixes in calculate_score via config model
                    # But we pass the raw string (e.g. "AM_2") so config can strip it
                    try:
                        current_date = pd.Timestamp(year=config.year, month=config.month, day=day_idx)

                        # Validate shift type existence (basic check)
                        base_type = get_base_shift_type(val)
                        if base_type in C.ACTIVE_DUTIES:
                            scaled_pts = config.points.calculate_score(
                                date_obj=current_date,
                                shift_type=val,  # Pass full string "AM_2"
                                scale=C.SCORE_SCALE_FACTOR,
                                holidays_obj=country_holidays,
                            )
                            roster_pts += scaled_pts / C.SCORE_SCALE_FACTOR
                    except Exception as e:
                        logger.warning(f"Error calculating score for {person} on {day_idx}: {e}")
                        continue

        # Month Pts = Points earned from duties + Manual Adjustments
        month_total = roster_pts + adj

        raw_total = bf + month_total

        summary.append(
            {
                "Name": person,
                "Brought Fwd": bf,
                "Roster Pts": roster_pts,  # New breakdown column
                "Manual Adj": adj,  # New breakdown column
                "Month Pts": month_total,
                "Raw Total": raw_total,
            }
        )
        raw_carry_overs.append(raw_total)

    min_carry = min(raw_carry_overs) if raw_carry_overs else 0.0
    final_stats = []
    for record in summary:
        # Normalize Carry Over so the lowest person starts at 0 next month
        record["Carry Over"] = record["Raw Total"] - min_carry
        final_stats.append(record)

    return pd.DataFrame(final_stats)


def export_to_excel_bytes(df_roster: pd.DataFrame, df_stats: pd.DataFrame, config: AppConfig) -> bytes:
    """Generates a downloadable Excel file (bytes)."""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = C.EXCEL_SHEET_TITLE

    header_days = [get_day_num(c) for c in df_roster.columns]
    ws.append(C.EXCEL_HEADERS_STATIC + header_days + C.EXCEL_HEADERS_SUFFIX)

    stats_map = df_stats.set_index("Name").to_dict("index")

    for person in sorted(config.personnel):
        row = [person]
        for d_col in df_roster.columns:
            val = df_roster.at[person, d_col] if person in df_roster.index else ""
            row.append(val)

        s = stats_map.get(person, {"Brought Fwd": 0, "Month Pts": 0, "Carry Over": 0})
        row.extend([s["Brought Fwd"], s["Month Pts"], s["Carry Over"]])
        ws.append(row)

    # Styles
    fill_header = PatternFill("solid", fgColor=C.COLOR_HEADER_BG.replace("#", ""))
    fill_x = PatternFill("solid", fgColor=C.COLOR_CONSTRAINT_BG.replace("#", ""))
    fill_24h = PatternFill("solid", fgColor=C.COLOR_FILL_24H)
    fill_am = PatternFill("solid", fgColor=C.COLOR_FILL_AM)
    fill_pm = PatternFill("solid", fgColor=C.COLOR_FILL_PM)
    fill_sb = PatternFill("solid", fgColor=C.COLOR_FILL_SB)

    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if cell.row == 1:
                cell.fill = fill_header
                cell.font = Font(bold=True)

            val_str = str(cell.value).upper() if cell.value else ""

            # Updated coloring logic to handle suffixes
            if val_str == "X":
                cell.fill = fill_x
            elif val_str.startswith("24H"):
                cell.fill = fill_24h
            elif val_str.startswith("AM"):
                cell.fill = fill_am
            elif val_str.startswith("PM"):
                cell.fill = fill_pm
            elif val_str.startswith("S/B"):
                cell.fill = fill_sb

    wb.save(output)
    return output.getvalue()
